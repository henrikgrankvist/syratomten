"""
Participants needs unique names
"""

DEBUG = True


from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
import datetime
import sys
import json
import copy

from module.googleapi import Google

NAME = 0
KLASS = 1
KLUBB = 2

dt = 3
workbook_sheets = ["Herr", "Dam", "Herr U23", "Dam U23"] # Optional improvement is to read the classes from the init workbook
workbooks_created = []
race_name = "Syratomten"
final_results_workbook_name = race_name + " Total Poängställning.xlsx"

google_sheet = {
        "spreadsheetId": "1bZAB1gelK82fzusrxsaYxEg2mntBaga6ds62nD8Eu1I",
        "sheetName": "Test",
        "range": "!A2:M"
}


column_dict1 = {
    0 : "A",
    1 : "B",
    2 : "C",
    3 : "D",
    4 : "E",
    5 : "F",
    6 : "G",
    7 : "H",
    8 : "I",
    9 : "J",
    10 : "K",
    11 : "L",
    12 : "M"
}

# This dictionary is translating the names of the races to a specific column in the final wourkbook.
column_dict = {
    "Syratomten Deltävling 1.xlsx" : "C",
    "Syratomten Deltävling 2.xlsx" : "D",
    "Syratomten Deltävling 3.xlsx" : "E",
    "Syratomten Deltävling 4.xlsx" : "F",
    "Syratomten Deltävling 5.xlsx" : "G",
    "Syratomten Deltävling 6.xlsx" : "H",
    "Syratomten Deltävling 7.xlsx" : "I",
    "Syratomten Deltävling 8.xlsx" : "J",
    "Syratomten Deltävling 9.xlsx" : "K",
    "Syratomten Deltävling Final.xlsx" : "L"
}


"""
Functions
"""

def scoreboard(name, klass, klubb, tid, number_of_participants, position):

    points = 5 + number_of_participants - position

    if tid != None: # If the tid is None, they have not raced
        pos1 = position + 1
        pos2 = str(position + 2)


        score_workbook[klass]["A" + pos2] = pos1
        score_workbook[klass]["A" + pos2].alignment = Alignment(horizontal='left')
        score_workbook[klass]["B" + pos2] = name # Name

        if klubb != None: # Dont write "None" as the club
            score_workbook[klass]["C" + pos2] = klubb # Klubb


        min2hrs = int(datetime.datetime.strptime(tid, "%M:%S").strftime("%M"))/60
        sec2hrs = int(datetime.datetime.strptime(tid, "%M:%S").strftime("%S"))/3600
        speed = 19.5 / (min2hrs + sec2hrs)

        score_workbook[klass]["D" + pos2] = tid  # Tid
        score_workbook[klass]["E" + pos2] = "{:.1f}".format(speed) # speed


        if klubb == "Väsby SS Triathlon": # Only Väsby Triathlon members gets a score
            score_workbook[klass]["F" + pos2] = points      # points
            score_workbook[klass]["F" + pos2].alignment = Alignment(horizontal='left')

        return pos1
    else: # If tid was none, dont increase the position
        return position

def fill_final_results():

    # Load the final score workbook
    final_workbook = load_workbook(filename=final_results_workbook_name)

    final_result_dict = {}

    for races in workbooks_created:

        # Load the individual race score workbook
        race_workbook = load_workbook(filename=races)
        print("INFO: Opened workbook " + races)

        # For each class in the race
        for race_class in workbook_sheets:

            # Append the workbook values to a dictionary
            for values in race_workbook[race_class].iter_rows(min_row=2, values_only=True):

                if race_class not in final_result_dict: # If the class doesn't exist in the dictionary, create the class
                    final_result_dict[race_class] = {}

                if values[5] != None: # Dont save the result if they dont have a score
                    if values[1] not in final_result_dict[race_class]: # if the participant doesn't exists
                        final_result_dict[race_class][values[1]] = {races: values[5]}
                    else:
                        final_result_dict[race_class][values[1]][races] = values[5]



    """
    This loop print the results saved in the dictionary to the final excel workbook.
    The results won't be sorted. That is fixed later.
    """
    for klass in final_result_dict:
        for row, name in enumerate(final_result_dict[klass], 2):
            total_points = 0
            for race in final_result_dict[klass][name]:

                # The dict translates the race name to a specific column
                column = column_dict[race]

                final_workbook[klass]["A" + str(row)].alignment = Alignment(horizontal='left') # Align the position

                final_workbook[klass]["B" + str(row)] = name

                final_workbook[klass][column + str(row)] = final_result_dict[klass][name][race] # The race score for each race
                final_workbook[klass][column + str(row)].alignment = Alignment(horizontal='center') # Align the race score

                total_points += final_result_dict[klass][name][race]

                #final_workbook[klass][column + str(row)].alignment = Alignment(horizontal='center')

            final_workbook[klass]["M" + str(row)] = total_points
            final_workbook[klass]["M" + str(row)].alignment = Alignment(horizontal='left')


    # Save the final_workbook after all the results are saved
    final_workbook.save(filename=final_results_workbook_name)
    print("INFO: The workbook " + final_results_workbook_name + " was saved.")





def sort_individual_race(elem):
    if elem[dt]: # If the values is not None, return that value
        return elem[dt]
    else: # If the value is None return "00:00" instead. Otherwise the sort() function will try to sort None, which doesn't work
        return "00:00"

def new_sort_individual_race(elem):
    if elem[2]: # If the values is not None, return that value
        return elem[2]
    else: # If the value is None return "00:00" instead. Otherwise the sort() function will try to sort None, which doesn't work
        return "00:00"

def sort_final_score(elem):
    return elem[12]

def create_race_workbook(workbook_name):

    # Create a workbook for each race
    workbook = Workbook()
    workbook.save(filename=race_name + " " + workbook_name + ".xlsx")
    print("INFO: Workbook " + race_name + " " + workbook_name + ".xlsx was created.")

    # Save the name of the workbooks created so I can open them later for the final results
    workbooks_created.append(race_name + " "  + workbook_name + ".xlsx")

    score_workbook = load_workbook(filename=race_name + " "  + workbook_name + ".xlsx")

    # Populate the workbook with the sheets listed in the list "workbook_sheets"
    for sheet in workbook_sheets:
        if sheet not in workbook.sheetnames:
            workbook.create_sheet(sheet)

            workbook[sheet]["A1"] = "Placering"
            workbook[sheet]["B1"] = "Namn"
            workbook[sheet]["C1"] = "Klubb"
            workbook[sheet]["D1"] = "Tid"
            workbook[sheet]["E1"] = "Hastighet (km/h)"
            workbook[sheet]["F1"] = "Poäng"

            # Setting column width in the workbook sheets
            workbook[sheet].column_dimensions["A"].width = 9
            workbook[sheet].column_dimensions["B"].width = 20
            workbook[sheet].column_dimensions["C"].width = 20
            workbook[sheet].column_dimensions["D"].width = 9
            workbook[sheet].column_dimensions["E"].width = 16
            workbook[sheet].column_dimensions["F"].width = 9

            print("INFO: Sheet " + sheet + " created in workbook " + str(workbook_name))

    # Remove the sheet named "Sheet", which is created by default.
    if "Sheet" in score_workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    return workbook

def create_final_results_workbook():

    # Create the workbook
    workbook = Workbook()

    workbook.save(filename=final_results_workbook_name)
    print("INFO: Workbook " + final_results_workbook_name + " was created.")

    score_workbook = load_workbook(filename=final_results_workbook_name)

    # Populate the workbook with the sheets listed in the list "workbook_sheets"
    for sheet in workbook_sheets:
        if sheet not in workbook.sheetnames:
            workbook.create_sheet(sheet)

            workbook[sheet]["A1"] = "Placering"
            workbook[sheet]["B1"] = "Namn"
            workbook[sheet]["C1"] = "DT1"
            workbook[sheet]["D1"] = "DT2"
            workbook[sheet]["E1"] = "DT3"
            workbook[sheet]["F1"] = "DT4"
            workbook[sheet]["G1"] = "DT5"
            workbook[sheet]["H1"] = "DT6"
            workbook[sheet]["I1"] = "DT7"
            workbook[sheet]["J1"] = "DT8"
            workbook[sheet]["K1"] = "DT9"
            workbook[sheet]["L1"] = "F"
            workbook[sheet]["M1"] = "Totalt"

            # Setting the column width
            workbook[sheet].column_dimensions["A"].width = 9
            workbook[sheet].column_dimensions["B"].width = 20
            workbook[sheet].column_dimensions["C"].width = 4
            workbook[sheet].column_dimensions["D"].width = 4
            workbook[sheet].column_dimensions["E"].width = 4
            workbook[sheet].column_dimensions["F"].width = 4
            workbook[sheet].column_dimensions["G"].width = 4
            workbook[sheet].column_dimensions["H"].width = 4
            workbook[sheet].column_dimensions["I"].width = 4
            workbook[sheet].column_dimensions["J"].width = 4
            workbook[sheet].column_dimensions["K"].width = 4
            workbook[sheet].column_dimensions["L"].width = 4
            workbook[sheet].column_dimensions["M"].width = 6

            print("INFO: Sheet " + sheet + " created in " + final_results_workbook_name)

    # Remove the sheet named "Sheet", which is created by default.
    if "Sheet" in score_workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    workbook.save(filename=final_results_workbook_name)
    print ("INFO: Workbook " + final_results_workbook_name + " was saved")

def sort_final_results():
    # Load the workbook that includes all the race results
    real_final_workbook = load_workbook(filename=final_results_workbook_name)
    print("INFO: Opened " + final_results_workbook_name + " to sort it.")

    for sheet in workbook_sheets:

        # Append all the values in the initial workbook to a list. It is easyier to work with
        final_results_list = []
        for values in real_final_workbook[sheet].iter_rows(min_row=2, values_only=True):
            final_results_list.append(values)

        # For each race in the workbook
        for race in real_final_workbook[sheet].iter_rows(min_row=1, max_row=1, min_col=4, values_only=True):

            for workbooks in race:

                # Sort the result_list based on the times
                try:
                    final_results_list.sort(key=sort_final_score, reverse = True)

                except TypeError:
                    print("WARNING: DID NOT SORT " + str(workbooks))

        for position,stuff in enumerate(final_results_list,1): #enumerate starts at 1
            for idx, items in enumerate(stuff):
                column = column_dict1[idx]
                real_final_workbook[sheet][str(column) + str(position+1)] = items
                real_final_workbook[sheet]["A" + str(position+1)] = position

    real_final_workbook.save(filename=final_results_workbook_name)
    print("INFO: The workbook " + final_results_workbook_name + " was saved.")

def calculate_speed(time):
    min2hrs = int(datetime.datetime.strptime(time, "%M:%S").strftime("%M"))/60
    sec2hrs = int(datetime.datetime.strptime(time, "%M:%S").strftime("%S"))/3600
    return 19.5 / (min2hrs + sec2hrs)

if __name__ == "__main__":


    # Load the workbook that includes all the race results
    init_workbook = load_workbook(filename="st-test2.xlsx")

    #main_workbook = Google.get(google_sheet["spreadsheetId"], google_sheet["sheetName"], google_sheet["range"])
    #races = main_workbook[0][3:] # Saves the race names based on the heading in the spreadsheet.
    races = ['Deltävling 1', 'Deltävling 2', 'Deltävling 3', 'Deltävling 4', 'Deltävling 5', 'Deltävling 6', 'Deltävling 7', 'Deltävling 8', 'Deltävling 9', 'Final']
    classes = ["Herr", "Dam"]

    """
    ##############################################
    """
    main_workbook = Google.get(google_sheet["spreadsheetId"], google_sheet["sheetName"], "!A2:M")


    """
    translate_idx_to_race = {
        3 : "Deltävling 1",
        4 : "Deltävling 2",
        5 : "Deltävling 3",
        6 : "Deltävling 4",
        7 : "Deltävling 5",
        8 : "Deltävling 6",
        9 : "Deltävling 7",
        10 : "Deltävling 8",
        11 : "Deltävling 9",
        12 : "Final",
    }
    """

    """
    Only cells that have data are returned. This breaks the script, so the script
    starts with filling blank cells up to column 12.
    """
    for participant in main_workbook:
        while len(participant) <= 12:
            participant.append("")


    """
    Takes the main_workbook which is a array and rebuilds it to an dictionary
    with the data more structured. Like this:
    {
        "race": "Deltävling 1"
        "participants" [
            {
                "name": "Henrik Grankvist",
                "class": "Herr",
                "club": "Väsby SS Triathlon",
                "result": 29:00
            },
            {
                "name": "Pontus Bohlin",
                "class": "Herr",
                "club": "Väsby SS Triathlon",
                "result": 29:40
            }
        ]
    }
    """
    total_race_result_list =  []
    for idx,race in enumerate(races):
        race_result_dict = {} # Without clearing the dictionary the data will be overwritten in the total_race_result_list
        race_result_dict["race"] = race
        race_result_dict["participants"] = []

        for participant in main_workbook:
            participant_structur = {} # Clear the dictionary
            if participant[3+idx]:
                participant_structur = {
                    "name": participant[0],
                    "class": participant[1],
                    "club": participant[2],
                    "result": participant[3+idx]
                }
                race_result_dict["participants"].append(participant_structur)


        total_race_result_list.append(race_result_dict)


    #print(json.dumps(total_race_result_list, sort_keys=False, indent=4))
    """
    To update a google spreadsheet the data needs to be a array.
    Each user and its result has its own list. The list is cleared between each
    race.
    """
    sheet_titles_list = [] # The sheet titles are sent as a list in the Google API

    # Fill the sheet_titles_list with the name of the sheets that the spreadsheet should have
    for race_class in classes:
        sheet_dict = {}
        sheet_dict["properties"] = {"title" : race_class}
        sheet_titles_list.append(sheet_dict)



    for race in total_race_result_list:

        #race_spreadsheet = Google.create_spreadsheet(race["race"], sheet_titles_list)
        #print(f'Created spreadsheet for {race["race"]}')

        herr_race_list = []
        dam_race_list = []

        herr_number_of_participants = 0
        dam_number_of_participants = 0


        for participant in race["participants"]:
            #print(participant["name"], participant["result"])
            update_user_list = []
            update_user_list.append(participant["name"])
            update_user_list.append(participant["club"])
            update_user_list.append(participant["result"])

            speed = calculate_speed(participant["result"])
            update_user_list.append("{:.1f}".format(speed))

            if participant["class"] == "Herr":
                herr_race_list.append(update_user_list)
                herr_number_of_participants += 1
            elif participant["class"] == "Dam":
                dam_race_list.append(update_user_list)
                dam_number_of_participants += 1


        if DEBUG: print(f"""Number of participants:
        Herr: {herr_number_of_participants}
        Dam: {dam_number_of_participants}""")

        if herr_number_of_participants == 0 and dam_number_of_participants == 0:
            continue # If there are not participants in the race, continue to the next race.

        race_spreadsheet = Google.create_spreadsheet("Poängräkning " + race["race"], sheet_titles_list)
        print(f'Created spreadsheet for {race["race"]}')

        race["spreadsheet_id"] = race_spreadsheet


        # Sort the lists based on the result
        herr_race_list.sort(key=new_sort_individual_race)
        dam_race_list.sort(key=new_sort_individual_race)

        position_herr = 0
        position_dam = 0

        for participant in herr_race_list:
            participant.insert(0, position_herr+1)
            if "Väsby SS Triathlon" in participant: # Lazy search for Väsby SS Traithlon as the club. Only members of Väsby SS Traithlon get points
                participant.append(5 + herr_number_of_participants - position_herr)
            position_herr +=1

        for participant in dam_race_list:
            participant.insert(0, position_dam+1)
            if "Väsby SS Triathlon" in participant: # Lazy search for Väsby SS Traithlon as the club. Only members of Väsby SS Traithlon get points
                participant.append(5 + dam_number_of_participants - position_dam)
            position_dam +=1



        for race_class in classes:
            if race_class == "Herr":
                Google.update(race_spreadsheet, race_class, google_sheet["range"], herr_race_list)
            elif race_class == "Dam":
                Google.update(race_spreadsheet, race_class, google_sheet["range"], dam_race_list)




    #print(json.dumps(total_race_result_list, sort_keys=False, indent=4))
    if DEBUG:
        for race in total_race_result_list:
            if "spreadsheet_id" in race:
                print(race["race"], race["spreadsheet_id"])







    """
    ##############################################
    """
    exit()
    #print(race_result_dict)
    for idx, participant in enumerate(main_workbook):
        print(str(idx), participant)
        if idx == 0:
            continue

        race_result_dict["Deltävling 1"][participant[1]][participant[0]] = participant[3]

    print(json.dumps(race_result_dict, sort_keys=False, indent=4))


    """
    ##############################################
    """
    main_workbook = Google.get(google_sheet["spreadsheetId"], google_sheet["sheetName"], google_sheet["range"])



    exit()
    for race in races:
        main_workbook.sort(key=sort_individual_race)
    race_result_list = []
    for participant in main_workbook:
        if participant[3]:
            #race_result_list.append(participant[0],participant[2],participant[3])
            print(participant[0],participant[3])

    race_spreadsheet = Google.create("Syratomten Deltävling 1")
    print(race_spreadsheet)
    Google.update(race_spreadsheet, "Blad1", google_sheet["range"], main_workbook)

    print("\n\n\n")
    exit()

    # Append all the values in the initial workbook to a list. It is easyier to work with
    result_list = []

    # The spreadsheet data is read as a tuple
    for values in init_workbook["Syra Tomten"].iter_rows(min_row=2, values_only=True):
        result_list.append(values)

    print(result_list)
    exit()

    #Google.create("Syratomten Deltävling 1")


    # For each race in the workbook
    for race in init_workbook["Syra Tomten"].iter_rows(min_row=1, max_row=1, min_col=4, values_only=True): # Reads the different race names in the spreadsheet

        for workbooks in race:


            # Sort the result_list based on the times
            try:
                result_list.sort(key=sort_individual_race)
                #print("SORTED " + str(workbooks))
            except TypeError:
                print("WARNING: DID NOT SORT " + str(workbooks))



            number_of_participants_herr = 0
            number_of_participants_dam = 0
            number_of_participants_herru23 = 0
            number_of_participants_damu23 = 0

            """
            Count the number of participants in each class in this race.
            The number of participant is used for setting the score
            """
            for participant in result_list:
                if participant[dt] != None: # Only if the time is not None, could probably match anything
                    if participant[KLASS] == "Herr" or participant[KLASS] == "Herr U23":
                        number_of_participants_herr = number_of_participants_herr + 1
                    elif participant[KLASS] == "Dam" or participant[KLASS] == "Dam U23":
                        number_of_participants_dam = number_of_participants_dam + 1

                    if participant[KLASS] == "Herr U23":
                        number_of_participants_herru23 = number_of_participants_herru23 + 1
                    elif participant[KLASS] == "Dam U23":
                        number_of_participants_damu23 = number_of_participants_damu23 + 1



            # Reset the participants position
            position_herr = 0
            position_dam = 0
            position_herru23 = 0
            position_damu23 = 0


            # Only continue if there are any participants in the race
            if number_of_participants_herr != 0 or number_of_participants_dam != 0 or number_of_participants_herru23 != 0 or number_of_participants_damu23 != 0:
            #if any(number_of_participants_herr, number_of_participants_dam, number_of_participants_herru23, number_of_participants_damu23) != 0:

                # Create a new workbook for this race
                score_workbook = create_race_workbook(workbooks)

                for stuff in result_list:

                    # Herr U23 is also counted in the Herr class
                    if stuff[KLASS] == "Herr" or stuff[KLASS] == "Herr U23":
                        position_herr = scoreboard(stuff[NAME], "Herr", stuff[KLUBB], stuff[dt], number_of_participants_herr, position_herr)

                    # Dam U23 is also counted in the Dam class
                    elif stuff[KLASS] == "Dam" or stuff[KLASS] == "Dam U23":
                        position_dam = scoreboard(stuff[NAME], "Dam", stuff[KLUBB], stuff[dt], number_of_participants_dam, position_dam)

                    if stuff[KLASS] == "Herr U23":
                        position_herru23 = scoreboard(stuff[NAME], stuff[KLASS], stuff[KLUBB], stuff[dt], number_of_participants_herru23, position_herru23)

                    elif stuff[KLASS] == "Dam U23":
                        position_damu23 = scoreboard(stuff[NAME], stuff[KLASS], stuff[KLUBB], stuff[dt], number_of_participants_damu23, position_damu23)

                # Increase the deltävling by one each loop
                dt += 1

                # Save the content in the score workbook
                score_workbook.save(filename=race_name + " "  + workbooks + ".xlsx")
                print ("INFO: Workbook " + race_name + " "  + workbooks + ".xlsx was saved")

    # Create the final results workbook
    create_final_results_workbook()

    # Fill the final results workbook
    fill_final_results()

    # Sort the results in the final results workbook
    sort_final_results()
